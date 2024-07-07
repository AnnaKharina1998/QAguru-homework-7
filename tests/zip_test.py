from zipfile import ZipFile
from script_os import PROJECT_DIR
from pypdf import PdfReader
from openpyxl import load_workbook
import os.path
import csv


def test_pdf():
    with ZipFile(os.path.join(PROJECT_DIR, "tmp_archive.zip"), 'r') as zf:
        with zf.open('example.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert 'PDF test file' in text

def test_xlsx():
    with ZipFile(os.path.join(PROJECT_DIR, "tmp_archive.zip"), 'r') as zip_file:
        with zip_file.open('example.xlsx') as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            value = sheet.cell(row=1, column=1).value
            name = 'United States'
            assert 'XLSX test file' in value

def test_csv():
    with ZipFile(os.path.join(PROJECT_DIR, "tmp_archive.zip"), 'r') as zip_file:
        with zip_file.open('example.csv', 'r') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csv_reader = list(csv.reader(content.splitlines()))

            assert 'CSV test file' in content
            assert csv_reader[0][0] == 'CSV test file'